# appendix/prototype/prototype_excel_exporter.py

```python
"""
prototype_excel_exporter.py

Export structured LLM accounting output to an Excel-compatible monthly journal.

This module does not call an LLM API.
It assumes that structured journal-entry candidates have already been generated
by an LLM, a manual ChatGPT workflow, or another process.

This is a Proof of Concept for the monthly journal layer described in the whitepaper.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from prototype_schema import JournalEntryCandidate


JOURNAL_HEADERS = [
    "Date",
    "Vendor",
    "Total Amount Including Tax",
    "Debit Account",
    "Debit Amount Excluding Tax",
    "Tax Type",
    "Tax Amount",
    "Credit Account",
    "Credit Amount",
    "Payment Method",
    "Payer",
    "Description",
    "Evidence Type",
    "Evidence Location",
    "audit_log",
    "Confirmation Status",
    "Notes",
]


class MonthlyJournalExporter:
    """
    Export journal-entry candidates to an Excel monthly journal.

    The monthly journal is an intermediate accounting dataset.
    It is not the final tax return.
    """

    def __init__(self, output_path: str | Path):
        self.output_path = Path(output_path)

    def export(self, entries: Iterable[JournalEntryCandidate]) -> Path:
        """
        Create or overwrite an Excel workbook with journal entries.

        Parameters
        ----------
        entries:
            Iterable of JournalEntryCandidate objects.

        Returns
        -------
        Path
            Path to the generated Excel file.
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Monthly Journal"

        self._write_headers(worksheet)

        for entry in entries:
            worksheet.append(self._entry_to_row(entry))

        self._adjust_column_widths(worksheet)

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(self.output_path)

        return self.output_path

    def append(self, entries: Iterable[JournalEntryCandidate]) -> Path:
        """
        Append journal entries to an existing workbook.

        If the workbook does not exist, it is created.
        """
        if self.output_path.exists():
            workbook = load_workbook(self.output_path)
            worksheet = workbook["Monthly Journal"] if "Monthly Journal" in workbook.sheetnames else workbook.active
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Monthly Journal"
            self._write_headers(worksheet)

        for entry in entries:
            worksheet.append(self._entry_to_row(entry))

        self._adjust_column_widths(worksheet)

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(self.output_path)

        return self.output_path

    @staticmethod
    def _write_headers(worksheet: Worksheet) -> None:
        worksheet.append(JOURNAL_HEADERS)

    @staticmethod
    def _entry_to_row(entry: JournalEntryCandidate) -> List[object]:
        return [
            entry.date,
            entry.vendor,
            entry.amount_tax_included,
            entry.debit_account,
            entry.debit_amount_tax_excluded,
            entry.tax_type,
            entry.tax_amount,
            entry.credit_account,
            entry.credit_amount,
            entry.payment_method,
            entry.payer,
            entry.description,
            entry.evidence_type,
            entry.evidence_location,
            entry.audit_log,
            entry.confirmation_status.value,
            entry.notes,
        ]

    @staticmethod
    def _adjust_column_widths(worksheet: Worksheet) -> None:
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def load_entries_from_json(path: str | Path) -> List[JournalEntryCandidate]:
    """
    Load journal-entry candidates from a JSON file.

    Supported format:
    [
      { ...entry... },
      { ...entry... }
    ]
    """
    path = Path(path)

    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)

    if not isinstance(data, list):
        raise ValueError("Input JSON must be a list of journal-entry objects.")

    return [JournalEntryCandidate.model_validate(item) for item in data]


def load_entries_from_jsonl(path: str | Path) -> List[JournalEntryCandidate]:
    """
    Load journal-entry candidates from a JSONL file.

    Supported format:
    one JSON object per line.
    """
    path = Path(path)
    entries: List[JournalEntryCandidate] = []

    with path.open("r", encoding="utf-8") as file:
        for line_number, line in enumerate(file, start=1):
            stripped = line.strip()
            if not stripped:
                continue

            try:
                item = json.loads(stripped)
                entries.append(JournalEntryCandidate.model_validate(item))
            except Exception as exc:
                raise ValueError(f"Invalid JSONL at line {line_number}: {exc}") from exc

    return entries


if __name__ == "__main__":
    sample_entries = [
        JournalEntryCandidate(
            date="2026-04-15",
            vendor="Example Electronics",
            amount_tax_included=55000,
            payment_method="Personal credit card",
            payer="Representative",
            debit_account="Supplies Expense",
            debit_amount_tax_excluded=50000,
            tax_type="Taxable purchase 10%",
            tax_amount=5000,
            credit_account="Officer Loan",
            credit_amount=55000,
            description="Business-use PC monitor purchase.",
            audit_log=(
                "The user states that the monitor is used for business development work. "
                "Payment was made by the representative's personal card, so the credit-side "
                "account is treated as an officer loan candidate. Final treatment should be reviewed."
            ),
            evidence_type="Paper receipt",
            evidence_location="2026_FiscalYear/04/01_original/",
            confirmation_status="confirmed",
            notes="Sample entry."
        )
    ]

    exporter = MonthlyJournalExporter("output/monthly_journal_sample.xlsx")
    result_path = exporter.export(sample_entries)

    print(f"Exported: {result_path}")
```
